"""Platform admin write workflows will live here."""

import csv
from io import StringIO, BytesIO
from datetime import datetime, date, time
import openpyxl
import xlrd
from django.utils import timezone

MONTH_MAP = {
    "jan": 1, "january": 1,
    "feb": 2, "february": 2,
    "mar": 3, "march": 3,
    "apr": 4, "april": 4,
    "may": 5,
    "jun": 6, "june": 6,
    "jul": 7, "july": 7,
    "aug": 8, "august": 8,
    "sep": 9, "september": 9, "sept": 9,
    "oct": 10, "october": 10,
    "nov": 11, "november": 11,
    "dec": 12, "december": 12
}

class LeapYearSkip(Exception):
    """Exception raised when a February 29 row is skipped on non-leap years."""
    pass

def is_leap_year(y: int) -> bool:
    return y % 4 == 0 and (y % 100 != 0 or y % 400 == 0)

def map_header(raw_header) -> str:
    if not raw_header:
        return ""
    h = str(raw_header).lower().strip().replace("_", "").replace("-", "").replace(" ", "")
    if h in ("date", "day", "dateofmonth", "dayofmonth"):
        return "day_or_date"
    if h in ("fajr", "fajar"):
        return "fajr"
    if h in ("sunrise", "shuruq"):
        return "sunrise"
    if h in ("dhuhr", "zuhr", "zuher", "zohar", "dohar"):
        return "dhuhr"
    if h in ("asr", "asar"):
        return "asr"
    if h in ("maghrib", "magrib", "maghreb"):
        return "maghrib"
    if h in ("isha", "esha", "ishaa", "isha'a"):
        return "isha"
    if h in ("city", "town"):
        return "city"
    if h in ("month", "mon"):
        return "month"
    if h in ("year", "yr"):
        return "year"
    return h

def has_required_headers(headers) -> bool:
    required = {"fajr", "sunrise", "dhuhr", "asr", "maghrib", "isha"}
    if not required.issubset(set(headers)):
        return False
    return "day_or_date" in headers or ("month" in headers and "day_or_date" in headers)

def parse_month(val) -> int:
    if not val:
        raise ValueError("Month value is missing")
    val_str = str(val).strip().lower()
    if val_str.isdigit():
        m = int(val_str)
        if 1 <= m <= 12:
            return m
        raise ValueError(f"Invalid month number: {m}")
    if val_str in MONTH_MAP:
        return MONTH_MAP[val_str]
    raise ValueError(f"Invalid month name: '{val_str}'")

def parse_day(val) -> int:
    if not val:
        raise ValueError("Day/Date value is missing")
    val_str = str(val).strip()
    if val_str.isdigit():
        return int(val_str)
    try:
        f = float(val_str)
        if f.is_integer():
            return int(f)
    except ValueError:
        pass
    raise ValueError(f"Invalid day value: '{val_str}'")

def parse_full_date(val, expected_year: int) -> tuple[date, int | None]:
    if not val:
        raise ValueError("Date value is missing")
    
    # 1. Native datetime / date / time object
    if isinstance(val, (datetime, date)):
        d_obj = val.date() if isinstance(val, datetime) else val
        original_year = d_obj.year
        if d_obj.month == 2 and d_obj.day == 29:
            if not is_leap_year(expected_year):
                raise LeapYearSkip()
            return date(expected_year, 2, 29), original_year
        return date(expected_year, d_obj.month, d_obj.day), original_year
        
    val_str = str(val).strip()
    
    # 2. Excel serial date
    try:
        f = float(val_str)
        if 1 <= f <= 1000000:
            serial = int(f)
            if serial == 60:
                d_obj = date(1900, 2, 28)
            elif serial < 60:
                d_obj = date(1899, 12, 31) + timezone.timedelta(days=serial)
            else:
                d_obj = date(1899, 12, 30) + timezone.timedelta(days=serial)
            
            original_year = d_obj.year
            if d_obj.month == 2 and d_obj.day == 29:
                if not is_leap_year(expected_year):
                    raise LeapYearSkip()
                return date(expected_year, 2, 29), original_year
            return date(expected_year, d_obj.month, d_obj.day), original_year
    except ValueError:
        pass
        
    # 3. String format parsing with year
    formats_with_year = [
        "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", 
        "%d-%b-%Y", "%d-%B-%Y", "%Y/%m/%d", "%d.%m.%Y"
    ]
    for fmt in formats_with_year:
        try:
            parsed = datetime.strptime(val_str, fmt).date()
            original_year = parsed.year
            if parsed.month == 2 and parsed.day == 29:
                if not is_leap_year(expected_year):
                    raise LeapYearSkip()
                return date(expected_year, 2, 29), original_year
            return date(expected_year, parsed.month, parsed.day), original_year
        except ValueError:
            pass
            
    # 4. String format parsing without year (using leap year 2000 as pivot to support 29-Feb)
    formats_without_year = [
        "%d-%b", "%b-%d", "%d/%m", "%m/%d", "%d-%B", "%B-%d"
    ]
    for fmt in formats_without_year:
        try:
            parsed = datetime.strptime(f"2000-{val_str}", f"%Y-{fmt}").date()
            if parsed.month == 2 and parsed.day == 29:
                if not is_leap_year(expected_year):
                    raise LeapYearSkip()
                return date(expected_year, 2, 29), None
            return date(expected_year, parsed.month, parsed.day), None
        except ValueError:
            pass
            
    raise ValueError(f"Unsupported date: '{val_str}'")

def parse_time(time_val) -> str:
    if not time_val:
        raise ValueError("Time value is missing")
    if isinstance(time_val, time):
        return time_val.strftime("%H:%M:%S")
    if isinstance(time_val, datetime):
        return time_val.time().strftime("%H:%M:%S")
    if hasattr(time_val, "hour") and hasattr(time_val, "minute"):
        s = getattr(time_val, "second", 0)
        return f"{time_val.hour:02d}:{time_val.minute:02d}:{s:02d}"

    time_str = str(time_val).strip().upper()
    
    # 1. Decimal fraction of day
    try:
        f = float(time_str)
        if 0.0 <= f <= 1.0:
            total_seconds = int(round(f * 24 * 3600))
            hours = total_seconds // 3600
            minutes = (total_seconds % 3600) // 60
            seconds = total_seconds % 60
            if hours >= 24:
                hours = 23; minutes = 59; seconds = 59
            return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
    except ValueError:
        pass
        
    # 2. String time formats
    formats = [
        "%H:%M", "%H:%M:%S", "%I:%M %p", "%I:%M:%S %p",
        "%H.%M", "%H.%M.%S"
    ]
    for fmt in formats:
        try:
            parsed = datetime.strptime(time_str, fmt).time()
            return parsed.strftime("%H:%M:%S")
        except ValueError:
            pass
            
    raise ValueError(f"Invalid format: '{time_str}'")


class TimetableParser:
    @staticmethod
    def parse_file(file_obj, expected_year: int, selected_city_name: str, sheet_name: str = None):
        """Parse a CSV or XLSX/XLS file containing prayer timetable records.
        
        Returns:
            (records, errors, warnings, sheet_names, active_sheet)
        """
        filename = file_obj.name.lower()
        if filename.endswith(".csv"):
            return TimetableParser._parse_csv(file_obj, expected_year, selected_city_name)
        elif filename.endswith(".xlsx"):
            return TimetableParser._parse_xlsx(file_obj, expected_year, selected_city_name, sheet_name)
        elif filename.endswith(".xls"):
            return TimetableParser._parse_xls(file_obj, expected_year, selected_city_name, sheet_name)
        else:
            raise ValueError("Unsupported file format. Please upload a CSV, XLS, or XLSX file.")

    @staticmethod
    def _parse_csv(file_obj, expected_year: int, selected_city_name: str):
        records = []
        errors = []
        warnings = []
        
        try:
            file_bytes = file_obj.read()
            try:
                decoded = file_bytes.decode('utf-8')
            except UnicodeDecodeError:
                decoded = file_bytes.decode('latin-1')
                
            io_string = StringIO(decoded)
            reader = csv.reader(io_string)
            rows = [r for r in reader]
        except Exception as e:
            raise ValueError(f"Failed to read CSV file: {str(e)}")
            
        if not rows:
            raise ValueError("CSV file is empty.")
            
        header_row_idx = -1
        raw_headers = []
        for idx, r in enumerate(rows):
            if any(cell.strip() for cell in r):
                header_row_idx = idx
                raw_headers = r
                break
                
        if header_row_idx == -1:
            raise ValueError("No header row found in CSV.")
            
        headers = [map_header(h) for h in raw_headers]
        if not has_required_headers(headers):
            required = {"fajr", "sunrise", "dhuhr", "asr", "maghrib", "isha"}
            raise ValueError(f"Missing required columns. Expected at least: {', '.join(required)}")
            
        for line_idx, r in enumerate(rows[header_row_idx + 1:], start=header_row_idx + 2):
            if not any(cell.strip() for cell in r):
                continue
                
            row_dict = dict(zip(headers, r))
            TimetableParser._process_row(row_dict, line_idx, expected_year, selected_city_name, records, errors, warnings)
            
        return records, errors, warnings, [], "Default"

    @staticmethod
    def _parse_xlsx(file_obj, expected_year: int, selected_city_name: str, sheet_name: str = None):
        records = []
        errors = []
        warnings = []
        
        try:
            wb = openpyxl.load_workbook(file_obj, data_only=True)
            sheet_names = wb.sheetnames
        except Exception as e:
            raise ValueError(f"Failed to load Excel workbook: {str(e)}")
            
        if not sheet_names:
            raise ValueError("Excel workbook has no sheets.")
            
        active_sheet = sheet_name
        if not active_sheet or active_sheet not in sheet_names:
            for name in sheet_names:
                sheet = wb[name]
                first_row = next(sheet.iter_rows(values_only=True), [])
                headers = [map_header(h) for h in first_row if h is not None]
                if has_required_headers(headers):
                    active_sheet = name
                    break
            if not active_sheet:
                active_sheet = sheet_names[0]
                
        sheet = wb[active_sheet]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            raise ValueError(f"Sheet '{active_sheet}' is empty.")
            
        header_row_idx = -1
        raw_headers = []
        for idx, r in enumerate(rows):
            if any(cell is not None and str(cell).strip() for cell in r):
                header_row_idx = idx
                raw_headers = r
                break
                
        if header_row_idx == -1:
            raise ValueError(f"No headers found in sheet '{active_sheet}'.")
            
        headers = [map_header(h) for h in raw_headers]
        if not has_required_headers(headers):
            required = {"fajr", "sunrise", "dhuhr", "asr", "maghrib", "isha"}
            raise ValueError(f"Sheet '{active_sheet}' missing required columns. Expected at least: {', '.join(required)}")
            
        for line_idx, r in enumerate(rows[header_row_idx + 1:], start=header_row_idx + 2):
            if not any(cell is not None and str(cell).strip() for cell in r):
                continue
                
            row_dict = dict(zip(headers, r))
            TimetableParser._process_row(row_dict, line_idx, expected_year, selected_city_name, records, errors, warnings)
            
        return records, errors, warnings, sheet_names, active_sheet

    @staticmethod
    def _parse_xls(file_obj, expected_year: int, selected_city_name: str, sheet_name: str = None):
        records = []
        errors = []
        warnings = []
        
        try:
            file_bytes = file_obj.read()
            wb = xlrd.open_workbook(file_contents=file_bytes)
            sheet_names = wb.sheet_names()
        except Exception as e:
            raise ValueError(f"Failed to load XLS workbook: {str(e)}")
            
        if not sheet_names:
            raise ValueError("XLS workbook has no sheets.")
            
        active_sheet = sheet_name
        if not active_sheet or active_sheet not in sheet_names:
            for name in sheet_names:
                sheet = wb.sheet_by_name(name)
                if sheet.nrows > 0:
                    first_row = [sheet.cell_value(0, col) for col in range(sheet.ncols)]
                    headers = [map_header(h) for h in first_row if h is not None]
                    if has_required_headers(headers):
                        active_sheet = name
                        break
            if not active_sheet:
                active_sheet = sheet_names[0]
                
        sheet = wb.sheet_by_name(active_sheet)
        if sheet.nrows == 0:
            raise ValueError(f"Sheet '{active_sheet}' is empty.")
            
        header_row_idx = -1
        raw_headers = []
        for r_idx in range(sheet.nrows):
            row_vals = [sheet.cell_value(r_idx, col) for col in range(sheet.ncols)]
            if any(cell is not None and str(cell).strip() for cell in row_vals):
                header_row_idx = r_idx
                raw_headers = row_vals
                break
                
        if header_row_idx == -1:
            raise ValueError(f"No headers found in sheet '{active_sheet}'.")
            
        headers = [map_header(h) for h in raw_headers]
        if not has_required_headers(headers):
            required = {"fajr", "sunrise", "dhuhr", "asr", "maghrib", "isha"}
            raise ValueError(f"Sheet '{active_sheet}' missing required columns. Expected at least: {', '.join(required)}")
            
        for r_idx in range(header_row_idx + 1, sheet.nrows):
            row_vals = [sheet.cell_value(r_idx, col) for col in range(sheet.ncols)]
            row_dict = {}
            for col_idx, h in enumerate(headers):
                cell = sheet.cell(r_idx, col_idx)
                val = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    val = datetime(*xlrd.xldate_as_tuple(val, wb.datemode))
                row_dict[h] = val
                
            if not any(v is not None and str(v).strip() for v in row_dict.values()):
                continue
                
            TimetableParser._process_row(row_dict, r_idx + 1, expected_year, selected_city_name, records, errors, warnings)
            
        return records, errors, warnings, sheet_names, active_sheet

    @staticmethod
    def _process_row(row, line_idx: int, expected_year: int, selected_city_name: str, records: list, errors: list, warnings: list):
        try:
            parsed_date = None
            original_year = None
            
            if "month" in row and "day_or_date" in row:
                try:
                    month_num = parse_month(row.get("month"))
                    day_num = parse_day(row.get("day_or_date"))
                    
                    if month_num == 2 and day_num == 29:
                        if not is_leap_year(expected_year):
                            raise LeapYearSkip()
                            
                    parsed_date = date(expected_year, month_num, day_num)
                    original_year = None
                except LeapYearSkip:
                    warnings.append(f"Row {line_idx}: Info - Skipped February 29 row since {expected_year} is not a leap year.")
                    return
                except ValueError as val_err:
                    errors.append(f"Row {line_idx}: Invalid Date components (Month/Day) - {str(val_err)}")
                    return
            else:
                date_val = row.get("day_or_date")
                try:
                    parsed_date, original_year = parse_full_date(date_val, expected_year)
                except LeapYearSkip:
                    warnings.append(f"Row {line_idx}: Info - Skipped February 29 row since {expected_year} is not a leap year.")
                    return
                except ValueError as val_err:
                    errors.append(f"Row {line_idx}: Invalid Date '{date_val}' - {str(val_err)}")
                    return
            
            if original_year is not None and original_year != expected_year:
                if original_year != 1900:
                    warnings.append(
                        f"Row {line_idx}: Date year {original_year} differs from selected upload year {expected_year}."
                    )
                
            if "city" in row and row.get("city"):
                file_city = str(row.get("city")).strip().lower().replace(" ", "")
                sel_city = selected_city_name.strip().lower().replace(" ", "")
                if file_city != sel_city:
                    warnings.append(
                        f"Row {line_idx}: City '{row.get('city')}' differs from selected city '{selected_city_name}'."
                    )
                    
            prayers = ["fajr", "sunrise", "dhuhr", "asr", "maghrib", "isha"]
            parsed_times = {}
            for prayer in prayers:
                time_val = row.get(prayer)
                try:
                    parsed_times[f"{prayer}_time"] = parse_time(time_val)
                except ValueError as val_err:
                    errors.append(f"Row {line_idx}: Invalid {prayer.capitalize()} time - {str(val_err)}")
                    
            if len(parsed_times) < len(prayers):
                return
                
            records.append({
                "date": parsed_date.strftime("%Y-%m-%d"),
                **parsed_times
            })
        except Exception as e:
            errors.append(f"Row {line_idx}: Unexpected processing error - {str(e)}")
